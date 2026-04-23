import sys
import os

output_lines = []

try:
    import pandas as pd
    all_sheets = pd.read_excel('D:/src/opportunity-tracker/OE-opportunities-tracker.xlsx', sheet_name=None, header=None)
    output_lines.append('SUCCESS with pandas')
    output_lines.append(f'SHEET NAMES: {list(all_sheets.keys())}')
    output_lines.append('')
    for sheet_name, df in all_sheets.items():
        output_lines.append(f'=== Sheet: {sheet_name} ===')
        output_lines.append(f'Shape: {df.shape}')
        output_lines.append('First 5 rows:')
        output_lines.append(df.head(5).to_string())
        output_lines.append('')
except Exception as e1:
    output_lines.append(f'pandas failed: {e1}')
    try:
        import openpyxl
        wb = openpyxl.load_workbook('D:/src/opportunity-tracker/OE-opportunities-tracker.xlsx', read_only=True, data_only=True)
        output_lines.append('SUCCESS with openpyxl')
        output_lines.append(f'SHEET NAMES: {wb.sheetnames}')
        output_lines.append('')
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output_lines.append(f'=== Sheet: {sheet_name} ===')
            rows = list(ws.iter_rows(values_only=True))
            for i, row in enumerate(rows[:5]):
                output_lines.append(f'Row {i}: {row}')
            output_lines.append('')
    except Exception as e2:
        output_lines.append(f'openpyxl also failed: {e2}')

with open('D:/src/opportunity-tracker/xl_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))
