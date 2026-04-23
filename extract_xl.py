import sys

output_lines = []

try:
    import pandas as pd
    all_sheets = pd.read_excel('D:/src/opportunity-tracker/OE-opportunities-tracker.xlsx', sheet_name=None, header=None)
    output_lines.append('SUCCESS with pandas')
    output_lines.append(f'SHEET NAMES: {list(all_sheets.keys())}')
    output_lines.append('')
    for sheet_name, df in all_sheets.items():
        output_lines.append(f'=== Sheet: {sheet_name} ===')
        output_lines.append(f'Total rows (including header): {len(df)}')
        output_lines.append(f'Total columns: {len(df.columns)}')
        output_lines.append('--- Row 0 (headers): ---')
        output_lines.append(str(list(df.iloc[0])))
        output_lines.append('--- Rows 1-5 (data): ---')
        for idx in range(1, min(6, len(df))):
            output_lines.append(f'Row {idx}: {list(df.iloc[idx])}')
        output_lines.append('')
except Exception as e1:
    output_lines.append(f'pandas failed: {e1}')
    try:
        import openpyxl
        wb = openpyxl.load_workbook('D:/src/opportunity-tracker/OE-opportunities-tracker.xlsx', read_only=True, data_only=True)
        output_lines.append('SUCCESS with openpyxl')
        output_lines.append(f'SHEET NAMES: {wb.sheetnames}')
        output_lines.append('')
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            output_lines.append(f'=== Sheet: {sheet_name} ===')
            output_lines.append(f'Total rows: {len(rows)}')
            output_lines.append('--- Row 0 (headers): ---')
            if rows:
                output_lines.append(str(list(rows[0])))
            output_lines.append('--- Rows 1-5 (data): ---')
            for i, row in enumerate(rows[1:6], 1):
                output_lines.append(f'Row {i}: {list(row)}')
            output_lines.append('')
    except Exception as e2:
        output_lines.append(f'openpyxl also failed: {e2}')

with open('D:/src/opportunity-tracker/xl_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))
