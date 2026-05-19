#!/usr/bin/env python3
"""
load_db.py — Clean and reload OE Opportunity Tracker database from Excel.

Reads OE-opportunities-tracker-data.xlsx, wipes all tables, then inserts:
  - Master data (deal stages, confidence levels, BUs, categories, subcategories,
    business categories, customers)
  - Users (admin + sales/presales from the Excel data)
  - Opportunities (all 51 data rows)

Usage:
    python load_db.py

Requirements:
    pip install openpyxl psycopg2-binary python-dotenv bcrypt
"""

import os
import sys
import uuid
import warnings
import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

# ── dependency check ──────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: pip install openpyxl")
try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    sys.exit("ERROR: pip install psycopg2-binary")
try:
    import bcrypt
except ImportError:
    sys.exit("ERROR: pip install bcrypt")
try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit("ERROR: pip install python-dotenv")

warnings.filterwarnings("ignore", category=UserWarning)

# ── paths ─────────────────────────────────────────────────────────────────────
ROOT  = Path(__file__).parent
EXCEL = ROOT / "OE-opportunities-tracker-data.xlsx"
ENV   = ROOT / "backend" / ".env"

# ── env ───────────────────────────────────────────────────────────────────────
if ENV.exists():
    load_dotenv(ENV)
else:
    load_dotenv()

DB_URL = os.environ.get("DATABASE_URL")
if not DB_URL:
    sys.exit(
        "ERROR: DATABASE_URL not set.\n"
        "Create backend/.env with:\n"
        "  DATABASE_URL=postgresql://user:pass@host:5432/dbname"
    )

# ── helpers ───────────────────────────────────────────────────────────────────
def new_id() -> str:
    return str(uuid.uuid4())

def clean(v):
    return v.strip() if isinstance(v, str) else v

def to_decimal(v, default="0") -> Decimal:
    if v is None:
        return Decimal(default)
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return Decimal(default)

def to_date(v):
    """Return datetime.date or None for non-date / text values (TBC, Lost, etc.)."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None

def hash_pw(plain: str) -> str:
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt(12)).decode()

# ── read workbook ─────────────────────────────────────────────────────────────
if not EXCEL.exists():
    sys.exit(f"ERROR: Excel file not found: {EXCEL}")

print(f"Reading {EXCEL.name} …")
wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

# ══ DEAL STAGES  (from "Opportunity stage" sheet) ═════════════════════════════
# Sheet layout: row 0 = title, row 1 = group labels, row 2 = column headers, rows 3+ = data
stage_sheet_rows = list(wb["Opportunity stage"].iter_rows(values_only=True))
DEAL_STAGES = []
for i, row in enumerate(stage_sheet_rows[3:], start=1):
    code, classification, status, prob = row[0], row[1], row[2], row[3]
    if not code:
        continue
    DEAL_STAGES.append({
        "id":                 new_id(),
        "code":               str(code).strip(),
        "classification":     str(classification).strip() if classification else "",
        "status":             str(status).strip() if status else "",
        "winning_probability": to_decimal(prob),
        "sort_order":         i,
    })

# ══ OE OPPORTUNITIES DATA  (from "OE Opportunities" sheet) ════════════════════
# Sheet layout: row 0 = title, row 1 = group headers, row 2 = column headers, rows 3+ = data
# Columns:
#   0  Sr. no.              8  PIN Presales          16 Unit price (INR)
#   1  Customer             9  Deal stage             17 Unit price (USD)
#   2  Description         10  Deal classification    18 TCV (USD Mln.)
#   3  Business Unit       11  Deal status            19 Comments
#   4  Product category    12  Winning probability    20 PMS
#   5  Product subcategory 13  Estimated closure date 21 Remarks
#   6  Business category   14  Confidence level
#   7  PIN Sales           15  Lifetime volume

opp_sheet_rows = list(wb["OE Opportunities"].iter_rows(values_only=True))
data_rows = [
    row for row in opp_sheet_rows[3:]
    if row[0] is not None and isinstance(row[0], (int, float))
    and row[1] is not None   # must have a customer
]

# ══ EXTRACT MASTER DATA from data rows + dropdown sheet ════════════════════════
customers_set   = set()
categories_set  = set()
subcat_to_cat   = {}   # subcat_name -> cat_name  (derived from actual data)
bizcats_set     = set()
bus_set         = set()
confidence_set  = set()
sales_names     = set()
presales_names  = set()

for row in data_rows:
    cust     = clean(row[1])
    bu       = clean(row[3])
    cat      = clean(row[4])
    subcat   = clean(row[5])
    bizcat   = clean(row[6])
    sales    = clean(row[7])
    presales = clean(row[8])
    conf     = clean(row[14])

    if cust:    customers_set.add(cust)
    if bu:      bus_set.add(bu)
    if cat:     categories_set.add(cat)
    if bizcat:  bizcats_set.add(bizcat)
    if conf:    confidence_set.add(conf)
    if sales:   sales_names.add(sales)
    if presales: presales_names.add(presales)
    if subcat and cat:
        existing = subcat_to_cat.get(subcat)
        if existing and existing != cat:
            print(f"  WARNING: subcategory '{subcat}' maps to both '{existing}' and '{cat}' — keeping first")
        else:
            subcat_to_cat[subcat] = cat

# Add all dropdown values as additional master data
dropdown_rows = list(wb["Drop down menu"].iter_rows(values_only=True))
# Column header row is index 2; data starts at index 3
# Columns: 0=Customer 1=ProductCat 2=Subcategory 3=BizCat 4=BizUnit 5=Stage 6=Closure 7=Confidence 8=Sales 9=Presales
for row in dropdown_rows[3:]:
    if clean(row[0]): customers_set.add(clean(row[0]))
    if clean(row[1]): categories_set.add(clean(row[1]))
    # subcategory-to-category mapping not derived from dropdown (no authoritative link)
    if clean(row[3]): bizcats_set.add(clean(row[3]))
    if clean(row[4]): bus_set.add(clean(row[4]))
    if clean(row[7]): confidence_set.add(clean(row[7]))
    if clean(row[8]): sales_names.add(clean(row[8]))
    if row[9] and clean(row[9]): presales_names.add(clean(row[9]))

# Remove None / empty strings that slipped through
customers_set  -= {None, ""}
categories_set -= {None, ""}
bizcats_set    -= {None, ""}
bus_set        -= {None, ""}
confidence_set -= {None, ""}
sales_names    -= {None, ""}
presales_names -= {None, ""}

# ══ CONFIDENCE LEVEL ORDER ════════════════════════════════════════════════════
CONF_ORDER = ["High", "Mid", "Low", "Secured", "Dropped", "On hold", "Lost"]
conf_ordered = [c for c in CONF_ORDER if c in confidence_set]
# append any that appear in data but aren't in our predefined order
for c in sorted(confidence_set):
    if c not in conf_ordered:
        conf_ordered.append(c)

# ══ USER DEFINITIONS ══════════════════════════════════════════════════════════
ADMIN_ID = new_id()

# Map display name → DB record
USER_DEFS = {
    "System Admin": {"id": ADMIN_ID,  "email": "admin@oe.local",   "role": "ADMIN",    "bu": None},
    "Manager":      {"id": new_id(),  "email": "manager@oe.local", "role": "MANAGER",  "bu": None},
    "Ashish":       {"id": new_id(),  "email": "ashish@oe.local",  "role": "SALES",    "bu": "MCC"},
    "Gaurav":       {"id": new_id(),  "email": "gaurav@oe.local",  "role": "SALES",    "bu": "MAC"},
    "Parvez":       {"id": new_id(),  "email": "parvez@oe.local",  "role": "SALES",    "bu": "MPC"},
    "KG/PA":        {"id": new_id(),  "email": "kg.pa@oe.local",   "role": "SALES",    "bu": None},
    "Rachit":       {"id": new_id(),  "email": "rachit@oe.local",  "role": "PRESALES", "bu": None},
    "Haseeb":       {"id": new_id(),  "email": "haseeb@oe.local",  "role": "PRESALES", "bu": None},
}

# ══ ID LOOKUP MAPS (populated during DB inserts) ══════════════════════════════
customer_ids = {}   # name  -> UUID
category_ids = {}   # name  -> UUID
subcat_ids   = {}   # name  -> UUID
bizcat_ids   = {}   # name  -> UUID
bu_ids       = {}   # name  -> UUID
stage_ids    = {}   # code  -> UUID
conf_ids     = {}   # name  -> UUID
user_ids     = {}   # name  -> UUID  (full_name -> id)


# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("Connecting to database …")
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur  = conn.cursor()

    try:
        # ── PHASE 1: CLEAN ────────────────────────────────────────────────────
        print("\n[1/3] Cleaning all tables …")
        cur.execute("""
            TRUNCATE
                audit_logs,
                notifications,
                document_access_logs,
                opportunity_history,
                documents,
                opportunities,
                users,
                customers,
                product_subcategories,
                product_categories,
                business_categories,
                business_units,
                deal_stages,
                confidence_levels
            RESTART IDENTITY CASCADE
        """)
        print("      ✓ All tables truncated")

        # ── PHASE 2: MASTER DATA ──────────────────────────────────────────────
        print("\n[2/3] Loading master data …")

        # Deal stages
        for s in DEAL_STAGES:
            cur.execute(
                """INSERT INTO deal_stages
                   (id, code, classification, status, winning_probability, sort_order, is_active)
                   VALUES (%s,%s,%s,%s,%s,%s,true)""",
                (s["id"], s["code"], s["classification"], s["status"],
                 s["winning_probability"], s["sort_order"]),
            )
            stage_ids[s["code"]] = s["id"]
        print(f"      ✓ {len(DEAL_STAGES)} deal stages")

        # Confidence levels
        for i, name in enumerate(conf_ordered, 1):
            cid = new_id()
            cur.execute(
                "INSERT INTO confidence_levels (id, name, sort_order) VALUES (%s,%s,%s)",
                (cid, name, i),
            )
            conf_ids[name] = cid
        print(f"      ✓ {len(conf_ordered)} confidence levels: {conf_ordered}")

        # Business units
        for name in sorted(bus_set):
            bid = new_id()
            cur.execute(
                "INSERT INTO business_units (id, name, is_active) VALUES (%s,%s,true)",
                (bid, name),
            )
            bu_ids[name] = bid
        print(f"      ✓ {len(bus_set)} business units: {sorted(bus_set)}")

        # Business categories
        for name in sorted(bizcats_set):
            bid = new_id()
            cur.execute(
                "INSERT INTO business_categories (id, name, is_active) VALUES (%s,%s,true)",
                (bid, name),
            )
            bizcat_ids[name] = bid
        print(f"      ✓ {len(bizcats_set)} business categories: {sorted(bizcats_set)}")

        # Product categories
        for name in sorted(categories_set):
            cid = new_id()
            cur.execute(
                "INSERT INTO product_categories (id, name, is_active) VALUES (%s,%s,true)",
                (cid, name),
            )
            category_ids[name] = cid
        print(f"      ✓ {len(categories_set)} product categories: {sorted(categories_set)}")

        # Product subcategories  (only those with a known category from the data)
        for subcat, cat in sorted(subcat_to_cat.items()):
            cat_id = category_ids.get(cat)
            if not cat_id:
                print(f"      WARNING: no category_id for '{cat}' (subcategory '{subcat}') — skipped")
                continue
            sid = new_id()
            cur.execute(
                """INSERT INTO product_subcategories (id, name, category_id, is_active)
                   VALUES (%s,%s,%s,true)""",
                (sid, subcat, cat_id),
            )
            subcat_ids[subcat] = sid
        print(f"      ✓ {len(subcat_ids)} product subcategories")

        # Customers
        for name in sorted(customers_set):
            cid = new_id()
            cur.execute(
                "INSERT INTO customers (id, name, is_active) VALUES (%s,%s,true)",
                (cid, name),
            )
            customer_ids[name] = cid
        print(f"      ✓ {len(customers_set)} customers")

        # Users
        print("      Hashing passwords (bcrypt 12 rounds) …")
        default_hash = hash_pw("Pioneer@123")
        for full_name, u in USER_DEFS.items():
            cur.execute(
                """INSERT INTO users
                   (id, full_name, email, password_hash, role, business_unit, is_active)
                   VALUES (%s,%s,%s,%s,%s,%s,true)""",
                (u["id"], full_name, u["email"], default_hash, u["role"], u["bu"]),
            )
            user_ids[full_name] = u["id"]
        print(f"      ✓ {len(USER_DEFS)} users  (default password: Pioneer@123)")

        # ── PHASE 3: OPPORTUNITIES ────────────────────────────────────────────
        print("\n[3/3] Inserting opportunities …")
        inserted = 0
        skipped  = 0

        for row in data_rows:
            sr_no    = int(row[0])
            cust     = clean(row[1])
            desc     = clean(row[2]) or "(no description)"
            bu       = clean(row[3])
            cat      = clean(row[4])
            subcat   = clean(row[5])
            bizcat   = clean(row[6])
            sales    = clean(row[7])
            presales = clean(row[8])
            stage    = clean(row[9])
            conf     = clean(row[14])
            lifetime = int(row[15]) if isinstance(row[15], (int, float)) else 0
            unit_inr = to_decimal(row[16])
            unit_usd = to_decimal(row[17])
            tcv      = to_decimal(row[18])
            comments = clean(row[19]) if row[19] else None
            pms      = clean(row[20]) if row[20] else None
            remarks  = clean(row[21]) if row[21] else None
            closure  = to_date(row[13])   # None for TBC / text values

            # Resolve foreign keys
            customer_id  = customer_ids.get(cust)
            bu_id        = bu_ids.get(bu)
            cat_id       = category_ids.get(cat)
            subcat_id    = subcat_ids.get(subcat)
            bizcat_id    = bizcat_ids.get(bizcat)
            stage_id     = stage_ids.get(stage)
            conf_id      = conf_ids.get(conf)
            sales_id     = user_ids.get(sales)
            presales_id  = user_ids.get(presales) if presales else None

            # Validate required FKs
            missing = []
            if not customer_id: missing.append(f'customer="{cust}"')
            if not bu_id:       missing.append(f'business_unit="{bu}"')
            if not cat_id:      missing.append(f'product_category="{cat}"')
            if not subcat_id:   missing.append(f'product_subcategory="{subcat}"')
            if not bizcat_id:   missing.append(f'business_category="{bizcat}"')
            if not stage_id:    missing.append(f'deal_stage="{stage}"')
            if not conf_id:     missing.append(f'confidence_level="{conf}"')
            if not sales_id:    missing.append(f'pin_sales="{sales}"')

            if missing:
                print(f"      SKIP row {sr_no:3d}: unresolved → {', '.join(missing)}")
                skipped += 1
                continue

            cur.execute(
                """INSERT INTO opportunities (
                       id, serial_number,
                       customer_id, description,
                       business_unit_id, product_category_id, product_subcategory_id,
                       business_category_id, pin_sales_id, pin_presales_id,
                       deal_stage_id, confidence_level_id,
                       estimated_closure_date, lifetime_volume,
                       unit_price_inr, unit_price_usd, tcv_usd_million,
                       comments, pms, remarks,
                       is_active, created_by, updated_by
                   ) VALUES (
                       %s,%s,
                       %s,%s,
                       %s,%s,%s,
                       %s,%s,%s,
                       %s,%s,
                       %s,%s,
                       %s,%s,%s,
                       %s,%s,%s,
                       true,%s,%s
                   )""",
                (
                    new_id(), sr_no,
                    customer_id, desc,
                    bu_id, cat_id, subcat_id,
                    bizcat_id, sales_id, presales_id,
                    stage_id, conf_id,
                    closure, lifetime,
                    unit_inr, unit_usd, tcv,
                    comments, pms, remarks,
                    ADMIN_ID, ADMIN_ID,
                ),
            )
            inserted += 1

        # Fix the serial_number auto-increment sequence so future inserts don't conflict
        cur.execute(
            """SELECT setval(
                   pg_get_serial_sequence('opportunities', 'serial_number'),
                   COALESCE((SELECT MAX(serial_number) FROM opportunities), 0),
                   true
               )"""
        )

        conn.commit()

        print(f"\n{'='*55}")
        print(f"  Done!  Inserted: {inserted}  |  Skipped: {skipped}")
        print(f"{'='*55}")
        if skipped:
            print("  (Skipped rows have unresolved references — check output above)")
        print(f"\n  Login credentials  (all users):  Pioneer@123")
        print(f"  Admin email  :  admin@oe.local")
        print(f"  Manager email:  manager@oe.local")

    except Exception as exc:
        conn.rollback()
        print(f"\nERROR — rolled back: {exc}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
